import openpyxl
from flask import Flask, render_template, request, redirect, url_for


# Load the Excel files
def load_excel(file_path):
    workbook = openpyxl.load_workbook(file_path)
    return workbook.active


# Load the Excel sheets
sheet = load_excel('ThaneFilteredEngg.xlsx')
mt5a_sheet = load_excel('data//MT5a.xlsx')
mt5m_sheet = load_excel('data//MT5m.xlsx')

app = Flask(__name__)


def find_info(query):
    if query.isdigit() and len(query) == 5:
        roll_no = 'P0' + query
        name = None

        # Search for the roll number in 'ThaneFilteredEngg.xlsx'
        for row in sheet.iter_rows(values_only=True):
            if row[0] == roll_no:
                name = row[1]
                break

        if name is None:
            return []

        # Search for the roll number in 'MT5a.xlsx'
        for mt5a_row in mt5a_sheet.iter_rows(values_only=True):
            if mt5a_row[0] == roll_no:
                marks_mt5a_paper1 = mt5a_row[1]
                marks_mt5a_paper2 = mt5a_row[2]
                total_marks_mt5a = mt5a_row[3]
                rank_mt5a = mt5a_row[4]
                break
        else:
            marks_mt5a_paper1 = "N/A"
            marks_mt5a_paper2 = "N/A"
            total_marks_mt5a = "N/A"
            rank_mt5a = "N/A"

        # Search for the roll number in 'MT5m.xlsx'
        for mt5m_row in mt5m_sheet.iter_rows(values_only=True):
            if mt5m_row[0] == roll_no:
                marks_mt5m = mt5m_row[1]
                rank_mt5m = mt5m_row[2]
                break
        else:
            marks_mt5m = "N/A"
            rank_mt5m = "N/A"

        return [{
            "name": name,
            "roll": roll_no,
            "marks_mt5a_paper1": marks_mt5a_paper1,
            "marks_mt5a_paper2": marks_mt5a_paper2,
            "total_marks_mt5a": total_marks_mt5a,
            "rank_mt5a": rank_mt5a,
            "marks_mt5m": marks_mt5m,
            "rank_mt5m": rank_mt5m
        }]

    else:
        matches = []
        for row in sheet.iter_rows(values_only=True):
            if query.lower() in row[1].lower():
                matches.append({"name": row[1], "roll": row[0]})
        return matches


@app.route('/')
def index():
    return render_template('index.html')


@app.route('/search', methods=['POST'])
def search():
    query = request.form['query']
    info = find_info(query)
    if info:
        return redirect(url_for('results', query=query))
    else:
        return "No matching records found"


@app.route('/results')
def results():
    query = request.args.get('query')
    info = find_info(query)
    return render_template('search_results.html', info=info)


if __name__ == '__main__':
    app.run(host='0.0.0.0', port=8080)
